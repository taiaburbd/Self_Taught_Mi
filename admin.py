# admin.py
import datetime
import sqlite3
from flask import Flask, render_template, request, redirect, url_for, session, flash
from app import app
from openpyxl import load_workbook
import secrets
import os
import nibabel as nib
import matplotlib.pyplot as plt


def is_admin():
    return session.get('user_type')

@app.before_request
def check_admin_access():
    if request.endpoint and request.endpoint.startswith('admin'):
        if not is_admin():
            return redirect(url_for('hello'))
            
def get_db_connection():
    conn = sqlite3.connect('database.db')
    conn.row_factory = sqlite3.Row
    return conn

#admin area start
@app.route('/admin')
def admin():
    conn = get_db_connection()
    posts = conn.execute('SELECT * FROM posts').fetchall()
    conn.close()
    return render_template('admin_index.html', posts=posts, utc_dt=datetime.datetime.utcnow())


@app.route('/admin_user_manage')
def admin_user_manager():
    conn = get_db_connection()
    users = conn.execute('SELECT * FROM users').fetchall()
    conn.close()

    return render_template('admin_user_manage.html', data=users)
#admin area end
@app.route('/admin_change_user_status/<int:user_id>/<int:user_status>', methods=['GET', 'POST'])
def admin_edit_user(user_id,user_status):
    conn = get_db_connection()
    users = conn.execute("UPDATE users SET isadmin = ? WHERE id = ?",
                       (user_status, user_id))
    conn.commit()
    conn.close()
    flash('User date successful updated! ','success')
    return redirect(url_for('admin_user_manager'))

@app.route('/admin_user_delete/<int:id>')
def delete_user(id):

    conn = get_db_connection()
    users = conn.execute("DELETE FROM users WHERE id = ?", (id,))
    conn.commit()
    conn.close()

    flash('User successful Removed! ','success')
    return redirect(url_for('admin_user_manager'))


#tutorial area start
@app.route('/tutorial_upload', methods=['POST'])
def upload():
    try:
        if 'file' not in request.files:
            return "No file part"
        file = request.files['file']
        if file.filename == '':
            return "No selected file"
        
        if file:
            conn = get_db_connection()
            cursor = conn.cursor()

            # Read the Excel file and insert data into the database
            workbook = load_workbook(file)
            sheet = workbook.active

            for row in sheet.iter_rows(values_only=True, min_row=2):
                cursor.execute('''
                    INSERT INTO tutorials (logo, title, description, details, content_type, level_id, category_id, status)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', row)

            conn.commit()
            conn.close()
        flash(f'Data successfully saved','success')
        return redirect(url_for('list_tutorials'))
    except Exception as e:
            flash(f'Try agian: {str(e)}','error')
            return redirect(url_for('list_tutorials'))
    
# Read operation (list all tutorials)
@app.route('/tutorial/list')
def list_tutorials():
    conn = get_db_connection()

    # Fetch data from the "level" table
    level_cursor = conn.execute('SELECT level_id, level_name FROM level')
    levels = level_cursor.fetchall()

    # Fetch data from the "category" table
    category_cursor = conn.execute('SELECT category_id, category_name FROM category')
    categories = category_cursor.fetchall()

    # Fetch data from the "tutorials" table
    tutorials_cursor = conn.execute('SELECT * FROM tutorials')
    tutorials = tutorials_cursor.fetchall()

    # Fetch data from the "summary" table
    level_count_cursor = conn.execute('SELECT level_id, COUNT(*) AS total_quantity FROM tutorials GROUP BY level_id')
    level_count = level_count_cursor.fetchall()

    cat_count_cursor = conn.execute('SELECT category_id, COUNT(*) AS total_quantity FROM tutorials GROUP BY category_id')
    cat_count = cat_count_cursor.fetchall()

    conn.close()

    return render_template('admin_list_tutorials.html', tutorials=tutorials, levels=levels, categories=categories, level_count=level_count,cat_count=cat_count)

# Delete tutorial
@app.route('/tutorial/delete/<int:tutorial_id>')
def delete_tutorial(tutorial_id):
    try:
        db = get_db_connection()
        cursor = db.cursor()
        cursor.execute("DELETE FROM tutorials WHERE tutorial_id = ?", (tutorial_id,))
        db.commit()
        cursor.close()
        flash(f'Data successfully deleted','success')
        return redirect(url_for('list_tutorials'))
    except Exception as e:
        flash(f'Try agian: {str(e)}','error')
        return redirect(url_for('list_tutorials'))

    return redirect(url_for('list_tutorials'))


#Question area start

@app.route('/question_upload', methods=['POST'])
def question_upload():
    try:
        if 'file' not in request.files:
            return "No file part"
        file = request.files['file']
        if file.filename == '':
            return "No selected file"
        
        if file:
            conn = get_db_connection()
            cursor = conn.cursor()

            # Read the Excel file and insert data into the database
            workbook = load_workbook(file)
            sheet = workbook.active

            for row in sheet.iter_rows(values_only=True, min_row=2):
                cursor.execute('''
                INSERT INTO questions (Question, Option1, Option2, Option3, Option4, Answer, Level, Category, Created_by)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', row)

            conn.commit()
            conn.close()
        flash(f'Data successfully saved','success')
        return redirect(url_for('list_questions'))
    except Exception as e:
            flash(f'Try agian: {str(e)}','error')
            return redirect(url_for('list_questions'))
    
# Read operation (list all tutorials)
@app.route('/questions/list')
def list_questions():
    conn = get_db_connection()

    # Fetch data from the "tutorials" table
    question_cursor = conn.execute('SELECT * FROM questions')
    list_questions = question_cursor.fetchall()

    # Fetch data from the "summary" table
    level_count_cursor = conn.execute('SELECT level, COUNT(*) AS total_quantity FROM questions GROUP BY level')
    level_count = level_count_cursor.fetchall()

    cat_count_cursor = conn.execute('SELECT category, COUNT(*) AS total_quantity FROM questions GROUP BY category')
    cat_count = cat_count_cursor.fetchall()

    conn.close()

    return render_template('admin_question.html', list_questions=list_questions, level_count=level_count,cat_count=cat_count)

# Delete 
@app.route('/question/delete/<int:question_id>')
def delete_question(question_id):
    try:
        db = get_db_connection()
        cursor = db.cursor()
        cursor.execute("DELETE FROM questions WHERE id = ?", (question_id,))
        db.commit()
        cursor.close()
        flash(f'Data successfully deleted','success')
        return redirect(url_for('list_questions'))
    except Exception as e:
        flash(f'Try agian: {str(e)}','error')
        return redirect(url_for('list_questions'))

#Question area end 


# Start Images aread
def save_slices_as_png(nifti_file_path):
    path = app.static_folder+"/uploads/"
    img = nib.load(path + nifti_file_path)
    data = img.get_fdata()

    # Create a directory to store the PNG files
    png_folder = path+ "png_slices/"
    os.makedirs(png_folder, exist_ok=True)

    print(nifti_file_path)

    name = os.path.splitext(nifti_file_path)[0]
    png_file_name = png_folder + name

    # Save each slice as a PNG file
    for i in range(data.shape[-1]):
        slice_data = data[:, :, i]
        plt.title(f'slice_{i}')
        plt.imshow(slice_data, cmap='gray')
        plt.axis('off')
        plt.savefig(f'{png_file_name}.png')
        plt.clf()


@app.route('/images_upload', methods=['POST'])
def images_upload():
    try:
        if 'file' not in request.files:
            return "No file part"
        file = request.files['file']
        if file.filename == '':
            return "No selected file"
        # Create the 'uploads' directory if it doesn't exist
        uploads_dir = app.static_folder+ '/uploads/'
        os.makedirs(uploads_dir, exist_ok=True)

        if file:
            conn = get_db_connection()

            category = request.form['category']
            title = request.form['title']

                # Generate a unique filename
            random_hex = secrets.token_hex(8)
            _, file_extension = os.path.splitext(file.filename)
            new_filename = random_hex + file_extension
            file_path = uploads_dir + new_filename
            png_file = random_hex + ".png"
            
            # Save the file
            file.save(file_path)

            save_slices_as_png(new_filename)

            sql = "INSERT INTO images_table (category, title, img_path,png_path, user_id, status) VALUES (?, ?, ?, ?, ?, ?)"
            conn.execute(sql,(category,title,new_filename,png_file,session['user_id'],'public'))
            
            conn.commit()
            conn.close()
        flash(f'Data successfully saved','success')
        return redirect(url_for('list_images'))
    except Exception as e:
            flash(f'Try agian: {str(e)}','error')
            return redirect(url_for('list_images'))
    
# Read (list all images)
@app.route('/images/list')
def list_images():
    conn = get_db_connection()

    # Fetch data from the "tutorials" table
    image_cursor = conn.execute('SELECT * FROM images_table')
    list_images = image_cursor.fetchall()

    # Fetch data from the "summary" table

    cat_count_cursor_questions = conn.execute('SELECT category_id, COUNT(*) AS total_quantity FROM tutorials GROUP BY category_id')
    cat_count_cursor_questions = cat_count_cursor_questions.fetchall()
    
    cat_count_cursor_images = conn.execute('SELECT category, COUNT(*) AS total_quantity FROM images_table GROUP BY category')
    cat_count_cursor_images = cat_count_cursor_images.fetchall()
    conn.close()

    return render_template('admin_images.html', list_images=list_images, cat_count_cursor_images=cat_count_cursor_images,cat_count_cursor_questions=cat_count_cursor_questions)

# Delete 
@app.route('/image_delete/<image_path>', methods=['GET'])
def delete_image(image_path):
    try:
        db = get_db_connection()
        cursor = db.cursor()
        image_id=param_value = request.args.get('image_id')

        cursor.execute("DELETE FROM images_table WHERE id = ?", (image_id))
        file_path = app.static_folder+ '/uploads/' + image_path
        name = os.path.splitext(image_path)[0]
        png_file_path = app.static_folder+ '/uploads/png_slices/' + name+'.png'

        if os.path.exists(file_path):
            os.remove(file_path)
        if os.path.exists(png_file_path):
            os.remove(png_file_path)

        db.commit()
        cursor.close()
        flash(f'Data successfully deleted','success')
        return redirect(url_for('list_images'))
    except Exception as e:
        flash(f'Try agian: {str(e)}','error')
        return redirect(url_for('list_images'))
# Images area end 
